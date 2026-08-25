from pathlib import Path
import os
import json
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.stats import rankdata, norm, t as tdist

ROOT = Path(os.environ.get("MICROBIOME_GATE_ROOT", Path(__file__).resolve().parents[1]))
XLSX = ROOT / "01_raw_data" / "crc_axis_external" / "yachida_2019" / "Yachida_Supplementary_Tables_1-15.xlsx"
BASE = ROOT / "06_results" / "crc_axis_external_yachida2019"
OUT = BASE / "independence_audit"
OUT.mkdir(parents=True, exist_ok=True)
LOCKED = ["Peptostreptococcus stomatis", "Gemella morbillorum", "Fusobacterium nucleatum", "Parvimonas unclassified"]
CRC_GROUPS = {"Stage_0", "Stage_I_II", "Stage_III_IV"}

def bh(p):
    p=np.asarray(p,float); n=len(p); o=np.argsort(p); z=p[o]*n/np.arange(1,n+1)
    z=np.minimum.accumulate(z[::-1])[::-1]; q=np.empty(n); q[o]=np.minimum(z,1); return q

def invnorm(x):
    x=np.asarray(x,float); ok=np.isfinite(x); out=np.full(len(x),np.nan)
    r=rankdata(x[ok],method="average"); out[ok]=norm.ppf((r-.5)/len(r)); return out

def matrix(ws, hr):
    data=list(ws.iter_rows(values_only=True)); h=list(data[hr-1]); valid=[0]+[i for i,v in enumerate(h) if i>0 and v is not None]
    return pd.DataFrame([[r[i] if i<len(r) else None for i in valid] for r in data[hr:] if r[0] is not None], columns=[str(h[i]) if h[i] is not None else "row_id" for i in valid])

def design(cov, species_df):
    C=pd.get_dummies(cov,drop_first=True,dtype=float).apply(pd.to_numeric,errors="coerce")
    C=C[[c for c in C if C[c].notna().sum()>=max(10,int(.75*len(C))) and C[c].nunique(dropna=True)>1]]
    S=pd.DataFrame({c:invnorm(species_df[c]) for c in species_df},index=species_df.index)
    return pd.concat([C,S],axis=1), list(S.columns)

def ols(y, X, target):
    y=np.asarray(y,float); ok=np.isfinite(y)&X.notna().all(axis=1).to_numpy(); D=np.column_stack([np.ones(ok.sum()),X.loc[ok].to_numpy(float)])
    if ok.sum() <= D.shape[1]+3: return (int(ok.sum()),np.nan,np.nan,np.nan)
    yy=y[ok]; b=np.linalg.lstsq(D,yy,rcond=None)[0]; resid=yy-D@b; df=len(yy)-D.shape[1]
    covb=np.linalg.pinv(D.T@D)*(resid@resid/df); j=1+X.columns.get_loc(target); se=np.sqrt(max(covb[j,j],0)); t=b[j]/se if se>0 else np.nan
    p=2*tdist.sf(abs(t),df) if np.isfinite(t) else np.nan
    return int(ok.sum()),float(b[j]),float(se),float(p)

wb=load_workbook(XLSX,read_only=True,data_only=True)
clinical=matrix(wb["Table_S2-2"],3).rename(columns={"Subject_ID":"subject_id"}); clinical.subject_id=clinical.subject_id.astype(str); clinical=clinical.set_index("subject_id")
met=matrix(wb["Table_S13"],4); met=met.rename(columns={met.columns[0]:"metabolite"}); M=met.set_index("metabolite").T.apply(pd.to_numeric,errors="coerce"); M.index=M.index.astype(str)
tax=matrix(wb["Table_S9"],4); tax=tax.rename(columns={tax.columns[0]:"species"}); A=tax.set_index("species").T.apply(pd.to_numeric,errors="coerce"); A.index=A.index.astype(str)
manifest=pd.read_csv(BASE/"participant_manifest.csv",dtype={"subject_id":str}).set_index("subject_id")
cons=pd.read_csv(BASE/"robustness"/"robustness_consensus.csv"); cand=cons[cons.strict_robust].copy()
ids=[i for i in manifest.index if i in clinical.index and i in M.index and i in A.index]
clinical=clinical.loc[ids]; M=M.loc[ids]; A=A.loc[ids]
pc=float(np.nanmin(A.to_numpy()[A.to_numpy()>0])/2); C=np.log(A+pc); C=C-C.mean(axis=1).to_numpy()[:,None]; S=C[LOCKED]

covcols=[c for c in ["Group","Age","Gender","BMI","Brinkman Index","Alcohol","Tumor location"] if c in clinical.columns]
crc_ids=clinical.index[clinical.Group.isin(CRC_GROUPS)]; healthy_ids=clinical.index[clinical.Group.eq("Healthy")]
crc_cov=[c for c in covcols if c not in {"Group"}]
if "Stage" in clinical.columns and "Stage" not in crc_cov: crc_cov.append("Stage")
healthy_cov=[c for c in covcols if c not in {"Group","Tumor location"}]

# Taxon correlation documents attribution ambiguity.
pd.DataFrame({c:invnorm(S[c]) for c in LOCKED},index=ids).corr(method="pearson").to_csv(OUT/"locked_species_rank_correlation.csv")

rows=[]
for _,a in cand.iterrows():
    sp,mt=a.species,a.metabolite
    if mt not in M.columns: continue
    y_all=invnorm(M.loc[ids,mt]); X_all,targets=design(clinical.loc[ids,covcols],S.loc[ids,LOCKED]); n,b,se,p=ols(y_all,X_all,sp)
    y_crc=invnorm(M.loc[crc_ids,mt]); X_crc,_=design(clinical.loc[crc_ids,crc_cov],S.loc[crc_ids,LOCKED]); nc,bc,sec,pcv=ols(y_crc,X_crc,sp)
    y_h=invnorm(M.loc[healthy_ids,mt]); X_h,_=design(clinical.loc[healthy_ids,healthy_cov],S.loc[healthy_ids,LOCKED]); nh,bh_,seh,ph=ols(y_h,X_h,sp)
    rows.append(dict(species=sp,metabolite=mt,n_all=n,beta_joint_all=b,se_joint_all=se,p_joint_all=p,n_crc=nc,beta_joint_crc=bc,se_joint_crc=sec,p_joint_crc=pcv,n_healthy=nh,beta_joint_healthy=bh_,se_joint_healthy=seh,p_joint_healthy=ph,
                     primary_discovery_r=a.partial_spearman_discovery,primary_validation_r=a.partial_spearman_validation,crc_univariable_r=a.crc_r,motu_r=a.motu_r,presence_r=a.presence_r))
R=pd.DataFrame(rows)
for col,qcol in [("p_joint_all","q_joint_all"),("p_joint_crc","q_joint_crc"),("p_joint_healthy","q_joint_healthy")]: R[qcol]=bh(R[col].fillna(1))
R["joint_all_concordant"]=np.sign(R.beta_joint_all)==np.sign(R.primary_discovery_r)
R["joint_crc_concordant"]=np.sign(R.beta_joint_crc)==np.sign(R.primary_discovery_r)
R["independent_crc_candidate"]=R.joint_all_concordant&(R.q_joint_all<.10)&R.joint_crc_concordant&(R.q_joint_crc<.10)
R["healthy_same_direction"]=np.sign(R.beta_joint_healthy)==np.sign(R.primary_discovery_r)
R["crc_preferential_exploratory"]=R.independent_crc_candidate & ((R.q_joint_healthy>=.10)|(~R.healthy_same_direction)) & (R.beta_joint_crc.abs()>R.beta_joint_healthy.abs())
R=R.sort_values(["crc_preferential_exploratory","independent_crc_candidate","q_joint_crc","q_joint_all"],ascending=[False,False,True,True])
R.to_csv(OUT/"joint_species_independence_results.csv",index=False)
top=R[R.crc_preferential_exploratory].head(20)
summary={"input_strict_robust_pairs":int(len(R)),"independent_crc_candidates":int(R.independent_crc_candidate.sum()),"crc_preferential_exploratory":int(R.crc_preferential_exploratory.sum()),"top_crc_preferential":top[["species","metabolite","beta_joint_all","q_joint_all","beta_joint_crc","q_joint_crc","beta_joint_healthy","q_joint_healthy"]].to_dict("records"),"interpretation_boundary":"Joint-species and CRC-preferential filters are post-validation attribution and prioritization audits, not new confirmatory gates; coefficients remain observational associations."}
(OUT/"independence_summary.json").write_text(json.dumps(summary,indent=2),encoding="utf-8")
print(json.dumps(summary,indent=2))
