from pathlib import Path
import os
import hashlib, json
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.stats import rankdata, norm, pearsonr

ROOT=Path(os.environ.get("MICROBIOME_GATE_ROOT", Path(__file__).resolve().parents[1]))
XLSX=ROOT/"01_raw_data"/"crc_axis_external"/"yachida_2019"/"Yachida_Supplementary_Tables_1-15.xlsx"
OUT=ROOT/"06_results"/"crc_axis_external_yachida2019"
OUT.mkdir(parents=True,exist_ok=True)
LOCKED=["Peptostreptococcus stomatis","Gemella morbillorum","Fusobacterium nucleatum","Parvimonas unclassified"]

def bh(p):
 p=np.asarray(p,float); n=len(p); o=np.argsort(p); z=p[o]*n/np.arange(1,n+1); z=np.minimum.accumulate(z[::-1])[::-1]; q=np.empty(n); q[o]=np.minimum(z,1); return q
def invnorm(x):
 x=np.asarray(x,float); ok=np.isfinite(x); out=np.full(len(x),np.nan); r=rankdata(x[ok],method="average"); out[ok]=norm.ppf((r-.5)/len(r)); return out
def residual(y,cov):
 X=pd.get_dummies(cov,drop_first=True,dtype=float).apply(pd.to_numeric,errors="coerce")
 keep=[c for c in X if X[c].notna().sum()>=max(10,int(.8*len(X))) and X[c].nunique(dropna=True)>1]
 X=X[keep]; y=np.asarray(y,float); ok=np.isfinite(y)&X.notna().all(axis=1).to_numpy(); D=np.column_stack([np.ones(ok.sum()),X.loc[ok].to_numpy(float)])
 b=np.linalg.lstsq(D,y[ok],rcond=None)[0]; out=np.full(len(y),np.nan); out[ok]=y[ok]-D@b; return out,keep
def sheet_matrix(ws, header_row=4):
 rows=ws.iter_rows(values_only=True); data=list(rows)
 hdr=list(data[header_row-1]); valid=[0]+[i for i,v in enumerate(hdr) if i>0 and v is not None]
 hdr=[str(hdr[i]) if hdr[i] is not None else "row_id" for i in valid]; vals=[]
 for row in data[header_row:]:
  if row[valid[0]] is None: continue
  vals.append([row[i] if i<len(row) else None for i in valid])
 return pd.DataFrame(vals,columns=hdr)

wb=load_workbook(XLSX,read_only=True,data_only=True)
clinical=sheet_matrix(wb["Table_S2-2"],3).rename(columns={"Subject_ID":"subject_id"})
species=sheet_matrix(wb["Table_S9"],4).rename(columns={"None":"feature"})
if species.columns[0] != "feature": species=species.rename(columns={species.columns[0]:"feature"})
metab=sheet_matrix(wb["Table_S13"],4)
metab=metab.rename(columns={metab.columns[0]:"metabolite"})

species=species[species.feature.astype(str).isin(LOCKED)].copy()
species_long=species.set_index("feature").T.apply(pd.to_numeric,errors="coerce")
species_long.index=species_long.index.astype(str); species_long.index.name="subject_id"
metab_long=metab.set_index("metabolite").T.apply(pd.to_numeric,errors="coerce")
metab_long.index=metab_long.index.astype(str); metab_long.index.name="subject_id"
clinical.subject_id=clinical.subject_id.astype(str)
ids=sorted(set(species_long.index)&set(metab_long.index)&set(clinical.subject_id))
clinical=clinical.set_index("subject_id").loc[ids]
S=species_long.loc[ids]; M=metab_long.loc[ids]

# Frozen deterministic 60/40 split within clinical Group using SHA-256 ordering.
split=pd.Series(index=ids,dtype=str)
for g,ix in clinical.groupby("Group").groups.items():
 ordered=sorted(list(ix),key=lambda s:hashlib.sha256(("YACHIDA_AXIS_V1|"+s).encode()).hexdigest())
 nd=max(1,int(np.floor(.60*len(ordered)))); split.loc[ordered[:nd]]="discovery"; split.loc[ordered[nd:]]="validation"

# CLR over the complete MetaPhlAn profile is required. Reload all rows, then subset matched subjects.
allsp=sheet_matrix(wb["Table_S9"],4); allsp=allsp.rename(columns={allsp.columns[0]:"feature"}); allsp["feature"]=allsp["feature"].astype(str).str.strip(); allsp=allsp.set_index("feature").T.apply(pd.to_numeric,errors="coerce")
allsp.index=allsp.index.astype(str); A=allsp.loc[ids].to_numpy(float); pos=A[A>0]; pc=float(np.nanmin(pos)/2)
clr=np.log(A+pc); clr=clr-clr.mean(axis=1,keepdims=True)
idx={f:i for i,f in enumerate(allsp.columns)}
Sclr=pd.DataFrame({f:clr[:,idx[f]] for f in LOCKED},index=ids)

covcols=[c for c in ["Group","Age","Gender","BMI","Brinkman Index","Alcohol","Tumor location"] if c in clinical.columns]
def run(part):
 ii=split[split.eq(part)].index; cov=clinical.loc[ii,covcols].copy(); rows=[]
 # Discovery-only metabolite filter: >=20% nonzero and nonconstant.
 if part=="discovery":
  keepm=[m for m in M if (pd.to_numeric(M.loc[ii,m],errors="coerce").fillna(0)!=0).mean()>=.20 and pd.to_numeric(M.loc[ii,m],errors="coerce").nunique(dropna=True)>2]
 else: keepm=list(M.columns)
 for s in LOCKED:
  xr,_=residual(invnorm(Sclr.loc[ii,s]),cov)
  for m in keepm:
   yr,_=residual(invnorm(pd.to_numeric(M.loc[ii,m],errors="coerce")),cov); ok=np.isfinite(xr)&np.isfinite(yr)
   if ok.sum()<25 or np.std(xr[ok])==0 or np.std(yr[ok])==0:r=p=np.nan
   else:r,p=pearsonr(xr[ok],yr[ok])
   rows.append(dict(partition=part,species=s,metabolite=m,n=int(ok.sum()),partial_spearman=float(r),p=float(p)))
 return pd.DataFrame(rows),keepm

disc,keepm=run("discovery"); disc["q_bh_family"]=bh(disc.p.fillna(1)); candidates=disc.loc[disc.q_bh_family<.10,["species","metabolite"]].drop_duplicates()
val_all,_=run("validation"); val=candidates.merge(val_all,on=["species","metabolite"],how="left")
if len(val): val["q_bh_locked"]=bh(val.p.fillna(1))
merged=disc.merge(val[["species","metabolite","n","partial_spearman","p","q_bh_locked"]] if len(val) else pd.DataFrame(columns=["species","metabolite","n","partial_spearman","p","q_bh_locked"]),on=["species","metabolite"],how="left",suffixes=("_discovery","_validation"))
merged["discovery_candidate"]=merged.q_bh_family<.10
merged["direction_concordant"]=np.sign(merged.partial_spearman_discovery)==np.sign(merged.partial_spearman_validation)
merged["heldout_reproduced"]=merged.discovery_candidate & merged.direction_concordant & (merged.q_bh_locked<.10)

detect=pd.DataFrame({"species":LOCKED,"prevalence_all":[float((S[s]>0).mean()) for s in LOCKED],"prevalence_discovery":[float((S.loc[split.eq('discovery'),s]>0).mean()) for s in LOCKED],"prevalence_validation":[float((S.loc[split.eq('validation'),s]>0).mean()) for s in LOCKED]})
manifest=clinical.copy(); manifest["partition"]=split; manifest.to_csv(OUT/"participant_manifest.csv")
detect.to_csv(OUT/"locked_species_detection.csv",index=False); disc.to_csv(OUT/"species_metabolite_discovery.csv",index=False); val.to_csv(OUT/"species_metabolite_locked_validation.csv",index=False); merged.to_csv(OUT/"species_metabolite_leakage_free_results.csv",index=False)
summary={"source_sha256":hashlib.sha256(XLSX.read_bytes()).hexdigest().upper(),"n_paired":len(ids),"n_discovery":int(split.eq('discovery').sum()),"n_validation":int(split.eq('validation').sum()),"groups":clinical.Group.value_counts().to_dict(),"locked_species":LOCKED,"n_metabolites_total":int(M.shape[1]),"n_metabolites_discovery_filtered":len(keepm),"n_tests_discovery":len(disc),"n_discovery_candidates":len(candidates),"n_heldout_reproduced":int(merged.heldout_reproduced.sum()),"reproduced_pairs":merged.loc[merged.heldout_reproduced,["species","metabolite","partial_spearman_discovery","partial_spearman_validation","q_bh_family","q_bh_locked"]].to_dict('records'),"claim_boundary":"Paired fecal metagenome-metabolome association after clinical adjustment; not microbial production, mediation, or causality."}
(OUT/"analysis_summary.json").write_text(json.dumps(summary,indent=2),encoding="utf-8"); print(json.dumps(summary,indent=2))
