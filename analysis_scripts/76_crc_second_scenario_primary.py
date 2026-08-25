from __future__ import annotations

import csv
import hashlib
import json
import math
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.stats import mannwhitneyu, rankdata, spearmanr
from statsmodels.stats.multitest import multipletests


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "01_raw_data" / "second_scenario_candidates" / "CRC_Yachida2019" / "41591_2019_458_MOESM3_ESM.xlsx"
OUT = ROOT / "06_results" / "second_scenario_crc"
OUT.mkdir(parents=True, exist_ok=True)

CRC_GROUPS = {"Stage_0", "Stage_I_II", "Stage_III_IV"}
PRIMARY_GROUPS = CRC_GROUPS | {"Healthy"}
DISCOVERY_FRACTION = 0.60
MIN_PREVALENCE = 0.10
PRIMARY_FDR = 0.05
VALIDATION_P = 0.05
PSEUDOCOUNT = 1e-6


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for block in iter(lambda: f.read(1024 * 1024), b""):
            h.update(block)
    return h.hexdigest()


def read_sheet_matrix(workbook, sheet: str, header_row: int = 4):
    ws = workbook[sheet]
    rows = ws.iter_rows(values_only=True)
    for _ in range(header_row - 1):
        next(rows)
    header = next(rows)
    ids = [str(x) for x in header[1:] if x is not None]
    values = []
    features = []
    width = len(ids) + 1
    for row in rows:
        row = row[:width]
        if not row or row[0] is None:
            continue
        features.append(str(row[0]))
        values.append([float(x) if x not in (None, "", "NA") else np.nan for x in row[1:]])
    return pd.DataFrame(np.asarray(values, dtype=float), index=features, columns=ids)


def bh(p):
    p = np.asarray(p, dtype=float)
    out = np.full(p.shape, np.nan)
    ok = np.isfinite(p)
    if ok.any():
        out[ok] = multipletests(p[ok], method="fdr_bh")[1]
    return out


def split_ids(meta: pd.DataFrame):
    split = {}
    for group, frame in meta.groupby("Group"):
        ordered = sorted(frame.index.astype(str), key=lambda x: hashlib.sha256(("CRC-v1|" + x).encode()).hexdigest())
        n_disc = max(1, min(len(ordered) - 1, round(len(ordered) * DISCOVERY_FRACTION)))
        for sid in ordered[:n_disc]:
            split[sid] = "discovery"
        for sid in ordered[n_disc:]:
            split[sid] = "validation"
    return pd.Series(split, name="split")


def feature_tests(matrix: pd.DataFrame, meta: pd.DataFrame, family: str, min_prevalence: float = MIN_PREVALENCE, discovery_fdr: float = PRIMARY_FDR):
    ids = [x for x in matrix.columns if x in meta.index]
    m = matrix[ids]
    md = meta.loc[ids]
    rows = []
    for feature, series in m.iterrows():
        x = pd.to_numeric(series, errors="coerce")
        prevalence = float(np.nanmean(np.asarray(x > 0, dtype=float)))
        if prevalence < min_prevalence:
            continue
        record = {"family": family, "feature": feature, "prevalence": prevalence}
        direction = {}
        for split in ("discovery", "validation", "all"):
            take = md.index if split == "all" else md.index[md["split"] == split]
            case = x[take][md.loc[take, "case"] == 1].dropna().to_numpy()
            ctrl = x[take][md.loc[take, "case"] == 0].dropna().to_numpy()
            if len(case) < 8 or len(ctrl) < 8:
                p = np.nan
                rb = np.nan
                log2fc = np.nan
            else:
                u, p = mannwhitneyu(case, ctrl, alternative="two-sided")
                rb = 2 * u / (len(case) * len(ctrl)) - 1
                log2fc = math.log2((np.nanmedian(case) + PSEUDOCOUNT) / (np.nanmedian(ctrl) + PSEUDOCOUNT))
            record.update({f"n_case_{split}": len(case), f"n_control_{split}": len(ctrl),
                           f"p_{split}": p, f"rank_biserial_{split}": rb, f"log2_median_ratio_{split}": log2fc})
            direction[split] = np.sign(rb) if np.isfinite(rb) else 0
        rows.append(record)
    result = pd.DataFrame(rows)
    if result.empty:
        return result
    result["q_discovery"] = bh(result["p_discovery"])
    result["q_validation_family"] = bh(result["p_validation"])
    result["q_all"] = bh(result["p_all"])
    result["reproduced"] = ((result["q_discovery"] < discovery_fdr) &
                             (result["p_validation"] < VALIDATION_P) &
                             (np.sign(result["rank_biserial_discovery"]) == np.sign(result["rank_biserial_validation"])) &
                             (result["q_all"] < PRIMARY_FDR))
    return result.sort_values(["reproduced", "q_all", "p_all"], ascending=[False, True, True])


def residualized_rank(values, meta):
    y = rankdata(values)
    age = pd.to_numeric(meta["Age"], errors="coerce").fillna(pd.to_numeric(meta["Age"], errors="coerce").median()).to_numpy()
    bmi = pd.to_numeric(meta["BMI"], errors="coerce").fillna(pd.to_numeric(meta["BMI"], errors="coerce").median()).to_numpy()
    sex = (meta["Gender"].astype(str).str.upper() == "M").astype(float).to_numpy()
    case = meta["case"].astype(float).to_numpy()
    X = np.column_stack([np.ones(len(meta)), case, rankdata(age), rankdata(bmi), sex])
    beta = np.linalg.lstsq(X, y, rcond=None)[0]
    return y - X @ beta


def link_tests(species, metabolites, meta, species_hits, metabolite_hits):
    rows = []
    for split in ("discovery", "validation"):
        ids = meta.index[meta["split"] == split]
        for sp in species_hits:
            for met in metabolite_hits:
                frame = pd.DataFrame({"sp": species.loc[sp, ids], "met": metabolites.loc[met, ids]}).dropna()
                if len(frame) < 20:
                    continue
                md = meta.loc[frame.index]
                rs = residualized_rank(np.log2(frame["sp"].to_numpy() + PSEUDOCOUNT), md)
                rm = residualized_rank(np.log2(frame["met"].to_numpy() + PSEUDOCOUNT), md)
                rho, p = spearmanr(rs, rm)
                rows.append({"species": sp, "metabolite": met, "split": split, "n": len(frame), "partial_spearman": rho, "p": p})
    links = pd.DataFrame(rows)
    if links.empty:
        return links
    discovery = links[links["split"] == "discovery"].copy()
    discovery["q_discovery"] = bh(discovery["p"])
    validation = links[links["split"] == "validation"].copy().rename(columns={"n": "n_validation", "partial_spearman": "rho_validation", "p": "p_validation"})
    merged = discovery.merge(validation[["species", "metabolite", "n_validation", "rho_validation", "p_validation"]], on=["species", "metabolite"], how="left")
    merged["reproduced_link"] = ((merged["q_discovery"] < PRIMARY_FDR) &
                                  (merged["p_validation"] < VALIDATION_P) &
                                  (np.sign(merged["partial_spearman"]) == np.sign(merged["rho_validation"])))
    return merged.sort_values(["reproduced_link", "q_discovery", "p_validation"], ascending=[False, True, True])


wb = load_workbook(SOURCE, read_only=True, data_only=True)
species = read_sheet_matrix(wb, "Table_S9")
ko = read_sheet_matrix(wb, "Table_S11")
metabolites = read_sheet_matrix(wb, "Table_S13")
wb.close()

meta = pd.read_excel(SOURCE, sheet_name="Table_S2-1", header=2)
meta["Subject_ID"] = meta["Subject_ID"].astype(str)
meta = meta.set_index("Subject_ID")
meta = meta[meta["Group"].isin(PRIMARY_GROUPS)].copy()
common = sorted(set(meta.index) & set(species.columns) & set(ko.columns) & set(metabolites.columns))
meta = meta.loc[common]
meta["case"] = meta["Group"].isin(CRC_GROUPS).astype(int)
meta["split"] = split_ids(meta)

species = species[common]
ko = ko[common]
metabolites = metabolites[common]

species_result = feature_tests(species, meta, "microbial_species")
ko_result = feature_tests(ko, meta, "microbial_KO")
metabolite_result = feature_tests(metabolites, meta, "fecal_metabolite")

species_hits = species_result.loc[species_result["reproduced"], "feature"].tolist()[:50]
metabolite_hits = metabolite_result.loc[metabolite_result["reproduced"], "feature"].tolist()[:50]
links = link_tests(species, metabolites, meta, species_hits, metabolite_hits)

primary_species_hits = set(species_hits)
sensitivity_rows = []
for prevalence_threshold in (0.05, 0.10, 0.20):
    for discovery_fdr in (0.025, 0.05, 0.10):
        alt = feature_tests(species, meta, "microbial_species", prevalence_threshold, discovery_fdr)
        hits = set(alt.loc[alt["reproduced"], "feature"])
        sensitivity_rows.append({
            "prevalence_threshold": prevalence_threshold,
            "discovery_fdr": discovery_fdr,
            "reproduced_species": len(hits),
            "primary_hits_retained": len(primary_species_hits & hits),
            "all_primary_hits_retained": primary_species_hits.issubset(hits),
            "hit_list": " | ".join(sorted(hits)),
        })
sensitivity = pd.DataFrame(sensitivity_rows)

meta.reset_index().to_csv(OUT / "participant_manifest.csv", index=False, encoding="utf-8-sig")
species_result.to_csv(OUT / "species_primary_results.csv", index=False, encoding="utf-8-sig")
ko_result.to_csv(OUT / "ko_primary_results.csv", index=False, encoding="utf-8-sig")
metabolite_result.to_csv(OUT / "metabolite_primary_results.csv", index=False, encoding="utf-8-sig")
links.to_csv(OUT / "species_metabolite_partial_links.csv", index=False, encoding="utf-8-sig")
sensitivity.to_csv(OUT / "species_threshold_sensitivity.csv", index=False, encoding="utf-8-sig")

summary = {
    "source": str(SOURCE.relative_to(ROOT)),
    "source_sha256": sha256(SOURCE),
    "primary_groups": sorted(PRIMARY_GROUPS),
    "n_participants_complete_three_matrix_overlap": len(meta),
    "group_counts": meta["Group"].value_counts().to_dict(),
    "split_counts": meta.groupby(["split", "Group"]).size().to_dict(),
    "features_tested": {"species": len(species_result), "KO": len(ko_result), "metabolites": len(metabolite_result)},
    "reproduced_features": {"species": int(species_result["reproduced"].sum()), "KO": int(ko_result["reproduced"].sum()), "metabolites": int(metabolite_result["reproduced"].sum())},
    "cross_layer_pairs_tested": len(links),
    "reproduced_cross_layer_links": int(links["reproduced_link"].sum()) if not links.empty else 0,
    "species_sensitivity_all_primary_retained_scenarios": int(sensitivity["all_primary_hits_retained"].sum()),
    "species_sensitivity_total_scenarios": len(sensitivity),
    "selection_rule": "outcome-blind disease selection; deterministic stratified discovery/validation split",
    "claim_boundary": "association and framework behaviour only; no causal or therapeutic claim"
}
summary["split_counts"] = {"|".join(k): int(v) for k, v in summary["split_counts"].items()}
(OUT / "primary_summary.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
print(json.dumps(summary, indent=2, ensure_ascii=False))
