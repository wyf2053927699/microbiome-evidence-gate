from __future__ import annotations

import hashlib
import json
import math
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.optimize import minimize
from scipy.special import expit
from scipy.stats import mannwhitneyu, norm

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "01_raw_data" / "second_scenario_candidates" / "CRC_Yachida2019" / "41591_2019_458_MOESM3_ESM.xlsx"
OUT = ROOT / "06_results" / "second_scenario_crc_v2"
OUT.mkdir(parents=True, exist_ok=True)

CRC_GROUPS = {"Stage_0", "Stage_I_II", "Stage_III_IV"}
PRIMARY_GROUPS = CRC_GROUPS | {"Healthy"}
DISCOVERY_FRACTION = 0.60
MIN_PREVALENCE = 0.10
DISCOVERY_FDR = 0.05
VALIDATION_FDR = 0.05
PSEUDOCOUNT = 1e-6
SEED = 20260819


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for block in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(block)
    return h.hexdigest()


def bh(values):
    p = np.asarray(values, dtype=float)
    q = np.full(p.shape, np.nan)
    ok = np.isfinite(p)
    if ok.any():
        pv = p[ok]
        order = np.argsort(pv)
        ranked = pv[order]
        adjusted = np.minimum.accumulate((ranked * len(ranked) / np.arange(1, len(ranked) + 1))[::-1])[::-1]
        restored = np.empty_like(adjusted)
        restored[order] = np.minimum(adjusted, 1.0)
        q[ok] = restored
    return q


def logistic_hc3(y, X):
    y = np.asarray(y, dtype=float)
    X = np.asarray(X, dtype=float)
    def objective(beta):
        eta = X @ beta
        return float(np.sum(np.logaddexp(0, eta) - y * eta))
    fit = minimize(objective, np.zeros(X.shape[1]), method="BFGS")
    if not fit.success:
        raise RuntimeError(fit.message)
    beta = fit.x
    mu = expit(X @ beta)
    w = np.clip(mu * (1 - mu), 1e-8, None)
    bread = np.linalg.pinv(X.T @ (w[:, None] * X))
    h = np.clip(w * np.einsum("ij,jk,ik->i", X, bread, X), 0, .999)
    scaled_score = X * ((y - mu) / (1 - h))[:, None]
    meat = scaled_score.T @ scaled_score
    cov = bread @ meat @ bread
    se = np.sqrt(np.clip(np.diag(cov), 0, None))
    z = beta / se
    p = 2 * norm.sf(np.abs(z))
    return beta, se, p


def read_sheet_matrix(workbook, sheet: str, header_row: int = 4):
    ws = workbook[sheet]
    rows = ws.iter_rows(values_only=True)
    for _ in range(header_row - 1):
        next(rows)
    header = next(rows)
    ids = [str(x) for x in header[1:] if x is not None]
    features, values = [], []
    for row in rows:
        row = row[: len(ids) + 1]
        if not row or row[0] is None:
            continue
        features.append(str(row[0]))
        values.append([float(x) if x not in (None, "", "NA") else np.nan for x in row[1:]])
    return pd.DataFrame(np.asarray(values, dtype=float), index=features, columns=ids)


def split_ids(meta: pd.DataFrame):
    split = {}
    for group, frame in meta.groupby("Group"):
        ordered = sorted(frame.index.astype(str), key=lambda x: hashlib.sha256(("CRC-v1|" + x).encode()).hexdigest())
        n_disc = max(1, min(len(ordered) - 1, round(len(ordered) * DISCOVERY_FRACTION)))
        split.update({sid: "discovery" for sid in ordered[:n_disc]})
        split.update({sid: "validation" for sid in ordered[n_disc:]})
    return pd.Series(split, name="split")


def rb_test(series: pd.Series, meta: pd.DataFrame, ids):
    ids = [x for x in ids if x in series.index]
    vals = pd.to_numeric(series.loc[ids], errors="coerce")
    md = meta.loc[ids]
    case = vals[md["case"] == 1].dropna().to_numpy()
    ctrl = vals[md["case"] == 0].dropna().to_numpy()
    if len(case) < 8 or len(ctrl) < 8:
        return len(case), len(ctrl), np.nan, np.nan, np.nan
    u, p = mannwhitneyu(case, ctrl, alternative="two-sided")
    rb = 2 * u / (len(case) * len(ctrl)) - 1
    med_case, med_ctrl = np.nanmedian(case), np.nanmedian(ctrl)
    l2 = math.log2((med_case + PSEUDOCOUNT) / (med_ctrl + PSEUDOCOUNT)) if med_case >= 0 and med_ctrl >= 0 else np.nan
    return len(case), len(ctrl), float(p), float(rb), float(l2)


def rb_bootstrap_ci(series: pd.Series, meta: pd.DataFrame, ids, rng, reps=2000):
    vals = pd.to_numeric(series.loc[ids], errors="coerce")
    md = meta.loc[ids]
    case = vals[md["case"] == 1].dropna().to_numpy()
    ctrl = vals[md["case"] == 0].dropna().to_numpy()
    if len(case) < 8 or len(ctrl) < 8:
        return np.nan, np.nan
    estimates = np.empty(reps)
    for i in range(reps):
        a = rng.choice(case, len(case), replace=True)
        b = rng.choice(ctrl, len(ctrl), replace=True)
        u = mannwhitneyu(a, b, alternative="two-sided").statistic
        estimates[i] = 2 * u / (len(a) * len(b)) - 1
    return tuple(np.quantile(estimates, [0.025, 0.975]))


def feature_family(matrix, meta, family, min_prevalence=MIN_PREVALENCE, discovery_fdr=DISCOVERY_FDR):
    disc_ids = meta.index[meta["split"] == "discovery"].tolist()
    val_ids = meta.index[meta["split"] == "validation"].tolist()
    all_ids = meta.index.tolist()
    rows = []
    for feature, series in matrix.iterrows():
        prevalence_discovery = float(np.nanmean((pd.to_numeric(series.loc[disc_ids], errors="coerce") > 0).astype(float)))
        if prevalence_discovery < min_prevalence:
            continue
        rec = {"family": family, "feature": feature, "prevalence_discovery": prevalence_discovery}
        for label, ids in (("discovery", disc_ids), ("validation", val_ids), ("all_descriptive", all_ids)):
            nc, nh, p, rb, l2 = rb_test(series, meta, ids)
            rec.update({f"n_case_{label}": nc, f"n_control_{label}": nh, f"p_{label}": p,
                        f"rank_biserial_{label}": rb, f"log2_median_ratio_{label}": l2})
        rows.append(rec)
    result = pd.DataFrame(rows)
    if result.empty:
        return result
    result["q_discovery_family"] = bh(result["p_discovery"])
    result["discovery_candidate"] = result["q_discovery_family"] < discovery_fdr
    result["q_validation_candidates"] = np.nan
    candidate_mask = result["discovery_candidate"]
    result.loc[candidate_mask, "q_validation_candidates"] = bh(result.loc[candidate_mask, "p_validation"])
    result["q_all_descriptive"] = bh(result["p_all_descriptive"])
    result["heldout_reproduced"] = (
        candidate_mask
        & (result["q_validation_candidates"] < VALIDATION_FDR)
        & (np.sign(result["rank_biserial_discovery"]) == np.sign(result["rank_biserial_validation"]))
    )
    rng = np.random.default_rng(SEED + {"microbial_species": 1, "microbial_KO": 2, "fecal_metabolite": 3}[family])
    for idx in result.index[result["discovery_candidate"]]:
        series = matrix.loc[result.loc[idx, "feature"]]
        for label, ids in (("discovery", disc_ids), ("validation", val_ids)):
            lo, hi = rb_bootstrap_ci(series, meta, ids, rng)
            result.loc[idx, f"rank_biserial_{label}_ci_low"] = lo
            result.loc[idx, f"rank_biserial_{label}_ci_high"] = hi
    return result.sort_values(["heldout_reproduced", "q_discovery_family", "p_validation"], ascending=[False, True, True])


def clr_matrix(matrix: pd.DataFrame):
    x = matrix.astype(float).clip(lower=0)
    logged = np.log(x + PSEUDOCOUNT)
    return logged.subtract(logged.mean(axis=0), axis=1)


def candidate_sensitivity(species, meta, primary):
    candidates = primary.loc[primary["discovery_candidate"], "feature"].tolist()
    disc_ids = meta.index[meta["split"] == "discovery"].tolist()
    val_ids = meta.index[meta["split"] == "validation"].tolist()
    clr = clr_matrix(species)
    rows = []
    for feature in candidates:
        rec = {"feature": feature}
        for label, ids in (("discovery", disc_ids), ("validation", val_ids)):
            _, _, p, rb, _ = rb_test(clr.loc[feature], meta, ids)
            rec.update({f"clr_p_{label}": p, f"clr_rank_biserial_{label}": rb})
        rows.append(rec)
    out = pd.DataFrame(rows)
    if not out.empty:
        out["clr_q_validation_candidates"] = bh(out["clr_p_validation"])
        out["clr_heldout_reproduced"] = ((out["clr_q_validation_candidates"] < VALIDATION_FDR) &
                                          (np.sign(out["clr_rank_biserial_discovery"]) == np.sign(out["clr_rank_biserial_validation"])))
    return out


def covariate_sensitivity(species, meta, primary):
    candidates = primary.loc[primary["discovery_candidate"], "feature"].tolist()
    clr = clr_matrix(species)
    rows = []
    for label in ("discovery", "validation"):
        ids = meta.index[meta["split"] == label]
        md = meta.loc[ids].copy()
        md["Age"] = pd.to_numeric(md["Age"], errors="coerce")
        md["BMI"] = pd.to_numeric(md["BMI"], errors="coerce")
        md["Brinkman Index"] = pd.to_numeric(md["Brinkman Index"], errors="coerce").mask(lambda x: x >= 990)
        md["Alcohol"] = pd.to_numeric(md["Alcohol"], errors="coerce").mask(lambda x: x >= 990)
        md["male"] = (md["Gender"].astype(str).str.upper() == "M").astype(float)
        for feature in candidates:
            frame = md[["case", "Age", "BMI", "Brinkman Index", "Alcohol", "male"]].copy()
            frame["feature_clr"] = clr.loc[feature, ids]
            for c in ["Age", "BMI", "Brinkman Index", "Alcohol"]:
                frame[c] = frame[c].fillna(frame[c].median())
            for c in ["Age", "BMI", "Brinkman Index", "Alcohol", "feature_clr"]:
                sd = frame[c].std(ddof=0)
                frame[c] = (frame[c] - frame[c].mean()) / sd if sd > 0 else 0.0
            cols = ["feature_clr", "Age", "BMI", "Brinkman Index", "Alcohol", "male"]
            X = np.column_stack([np.ones(len(frame)), frame[cols].to_numpy(dtype=float)])
            try:
                coef, ses, pvals = logistic_hc3(frame["case"].to_numpy(), X)
                beta, se, p = coef[1], ses[1], pvals[1]
                rows.append({"feature": feature, "split": label, "n": len(frame), "beta_feature_clr": beta,
                             "se_hc3": se, "ci_low": beta - 1.96 * se, "ci_high": beta + 1.96 * se, "p": p,
                             "covariates": "age + sex + BMI + Brinkman index + alcohol"})
            except Exception as exc:
                rows.append({"feature": feature, "split": label, "n": len(frame), "error": str(exc)})
    out = pd.DataFrame(rows)
    if not out.empty:
        out["q_within_split_candidates"] = np.nan
        for label in out["split"].unique():
            mask = out["split"] == label
            out.loc[mask, "q_within_split_candidates"] = bh(out.loc[mask, "p"])
    return out


def stage_descriptives(species, meta, primary):
    rows = []
    candidates = primary.loc[primary["discovery_candidate"], "feature"].tolist()
    for feature in candidates:
        for group, ids in meta.groupby("Group").groups.items():
            v = pd.to_numeric(species.loc[feature, list(ids)], errors="coerce")
            rows.append({"feature": feature, "group": group, "n": int(v.notna().sum()),
                         "prevalence": float((v > 0).mean()), "median": float(v.median()),
                         "q1": float(v.quantile(.25)), "q3": float(v.quantile(.75))})
    return pd.DataFrame(rows)


wb = load_workbook(SOURCE, read_only=True, data_only=True)
species = read_sheet_matrix(wb, "Table_S9")
ko = read_sheet_matrix(wb, "Table_S11")
metabolites = read_sheet_matrix(wb, "Table_S13")
wb.close()

meta = pd.read_excel(SOURCE, sheet_name="Table_S2-1", header=2)
meta["Subject_ID"] = meta["Subject_ID"].astype(str)
meta = meta.set_index("Subject_ID")
meta = meta[meta["Group"].isin(PRIMARY_GROUPS)].copy()
common = sorted(set(meta.index) & set(species.columns) & set(ko.columns) & set(metabolites.columns))
meta = meta.loc[common]
meta["case"] = meta["Group"].isin(CRC_GROUPS).astype(int)
meta["split"] = split_ids(meta)
species, ko, metabolites = species[common], ko[common], metabolites[common]

species_result = feature_family(species, meta, "microbial_species")
ko_result = feature_family(ko, meta, "microbial_KO")
metabolite_result = feature_family(metabolites, meta, "fecal_metabolite")
clr_sensitivity = candidate_sensitivity(species, meta, species_result)
covariate = covariate_sensitivity(species, meta, species_result)
stage = stage_descriptives(species, meta, species_result)

sensitivity_rows = []
primary_hits = set(species_result.loc[species_result["heldout_reproduced"], "feature"])
for prevalence in (0.05, 0.10, 0.20):
    for fdr in (0.025, 0.05, 0.10):
        alt = feature_family(species, meta, "microbial_species", prevalence, fdr)
        hits = set(alt.loc[alt["heldout_reproduced"], "feature"])
        sensitivity_rows.append({"prevalence_discovery": prevalence, "discovery_fdr": fdr,
                                 "discovery_candidates": int(alt["discovery_candidate"].sum()),
                                 "heldout_reproduced": len(hits), "primary_hits_retained": len(primary_hits & hits),
                                 "all_primary_hits_retained": primary_hits.issubset(hits), "hit_list": " | ".join(sorted(hits))})
sensitivity = pd.DataFrame(sensitivity_rows)

meta.reset_index().to_csv(OUT / "participant_manifest.csv", index=False, encoding="utf-8-sig")
species_result.to_csv(OUT / "species_leakage_free_results.csv", index=False, encoding="utf-8-sig")
ko_result.to_csv(OUT / "ko_leakage_free_results.csv", index=False, encoding="utf-8-sig")
metabolite_result.to_csv(OUT / "metabolite_leakage_free_results.csv", index=False, encoding="utf-8-sig")
clr_sensitivity.to_csv(OUT / "species_clr_sensitivity.csv", index=False, encoding="utf-8-sig")
covariate.to_csv(OUT / "species_covariate_sensitivity.csv", index=False, encoding="utf-8-sig")
stage.to_csv(OUT / "species_stage_descriptives.csv", index=False, encoding="utf-8-sig")
sensitivity.to_csv(OUT / "species_threshold_sensitivity.csv", index=False, encoding="utf-8-sig")

summary = {
    "analysis_version": "CRC leakage-free validation v2",
    "source": str(SOURCE.relative_to(ROOT)),
    "source_sha256": sha256(SOURCE),
    "independent_unit": "participant",
    "n_complete_overlap": len(meta),
    "split_counts": {"|".join(k): int(v) for k, v in meta.groupby(["split", "Group"]).size().to_dict().items()},
    "filtering": "discovery-only prevalence",
    "validation_family": "discovery candidates only",
    "validation_multiplicity": "Benjamini-Hochberg q < 0.05",
    "full_cohort_role": "descriptive only; excluded from replication decision",
    "feature_counts": {
        "species_eligible_discovery": len(species_result), "species_candidates": int(species_result["discovery_candidate"].sum()),
        "species_heldout_reproduced": int(species_result["heldout_reproduced"].sum()),
        "KO_eligible_discovery": len(ko_result), "KO_candidates": int(ko_result["discovery_candidate"].sum()),
        "KO_heldout_reproduced": int(ko_result["heldout_reproduced"].sum()),
        "metabolite_eligible_discovery": len(metabolite_result), "metabolite_candidates": int(metabolite_result["discovery_candidate"].sum()),
        "metabolite_heldout_reproduced": int(metabolite_result["heldout_reproduced"].sum()),
    },
    "species_hits": species_result.loc[species_result["heldout_reproduced"], "feature"].tolist(),
    "claim_boundary": "internal held-out consistency; not independent-cohort replication or causality",
}
(OUT / "analysis_summary.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
print(json.dumps(summary, indent=2, ensure_ascii=False))
