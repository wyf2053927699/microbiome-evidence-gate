from pathlib import Path
import os
import json, numpy as np, pandas as pd
from openpyxl import load_workbook
from scipy.stats import rankdata, norm, pearsonr

ROOT=Path(os.environ.get("MICROBIOME_GATE_ROOT", Path(__file__).resolve().parents[1]))
XLSX=ROOT/"01_raw_data"/"crc_axis_external"/"yachida_2019"/"Yachida_Supplementary_Tables_1-15.xlsx"
BASE=ROOT/"06_results"/"crc_axis_external_yachida2019"; OUT=BASE/"robustness"; OUT.mkdir(parents=True,exist_ok=True)
LOCKED=["Peptostreptococcus stomatis","Gemella morbillorum","Fusobacterium nucleatum","Parvimonas unclassified"]
CRC_GROUPS={"Stage_0","Stage_I_II","Stage_III_IV"}

def bh(p):
 p=np.asarray(p,float); n=len(p); o=np.argsort(p); z=p[o]*n/np.arange(1,n+1); z=np.minimum.accumulate(z[::-1])[::-1]; q=np.empty(n); q[o]=np.minimum(z,1); return q
def invnorm(x):
 x=np.asarray(x,float); ok=np.isfinite(x); out=np.full(len(x),np.nan); r=rankdata(x[ok],method="average"); out[ok]=norm.ppf((r-.5)/len(r)); return out
def residual(y,cov):
 X=pd.get_dummies(cov,drop_first=True,dtype=float).apply(pd.to_numeric,errors="coerce")
 keep=[c for c in X if X[c].notna().sum()>=max(10,int(.75*len(X))) and X[c].nunique(dropna=True)>1]; X=X[keep]; y=np.asarray(y,float)
 ok=np.isfinite(y)&X.notna().all(axis=1).to_numpy(); D=np.column_stack([np.ones(ok.sum()),X.loc[ok].to_numpy(float)]); b=np.linalg.lstsq(D,y[ok],rcond=None)[0]
 out=np.full(len(y),np.nan); out[ok]=y[ok]-D@b; return out
def matrix(ws,hr):
 data=list(ws.iter_rows(values_only=True)); h=list(data[hr-1]); valid=[0]+[i for i,v in enumerate(h) if i>0 and v is not None]; names=[str(h[i]) if h[i] is not None else "row_id" for i in valid]
 return pd.DataFrame([[r[i] if i<len(r) else None for i in valid] for r in data[hr:] if r[valid[0]] is not None],columns=names)

wb=load_workbook(XLSX,read_only=True,data_only=True)
clinical=matrix(wb["Table_S2-2"],3).rename(columns={"Subject_ID":"subject_id"}); clinical.subject_id=clinical.subject_id.astype(str); clinical=clinical.set_index("subject_id")
met=matrix(wb["Table_S13"],4); met=met.rename(columns={met.columns[0]:"metabolite"}); M=met.set_index("metabolite").T.apply(pd.to_numeric,errors="coerce"); M.index=M.index.astype(str)
manifest=pd.read_csv(BASE/"participant_manifest.csv",dtype={"subject_id":str}).set_index("subject_id")
primary=pd.read_csv(BASE/"species_metabolite_leakage_free_results.csv"); cand=primary[primary.heldout_reproduced].copy()
ids=[i for i in manifest.index if i in M.index]; clinical=clinical.loc[ids]; M=M.loc[ids]; manifest=manifest.loc[ids]
covcols=[c for c in ["Group","Age","Gender","BMI","Brinkman Index","Alcohol","Tumor location"] if c in clinical.columns]

def profile(sheet):
 d=matrix(wb[sheet],4); d=d.rename(columns={d.columns[0]:"feature"}); d.feature=d.feature.astype(str).str.strip(); X=d.set_index("feature").T.apply(pd.to_numeric,errors="coerce"); X.index=X.index.astype(str); X=X.loc[ids]
 A=X.to_numpy(float); pos=A[A>0]; pc=float(np.nanmin(pos)/2); C=np.log(A+pc); C=C-C.mean(axis=1,keepdims=True); return X,pd.DataFrame(C,index=ids,columns=X.columns),pc

meta_raw,meta_clr,meta_pc=profile("Table_S9"); motu_raw,motu_clr,motu_pc=profile("Table_S8-1")
motu_map={"Peptostreptococcus stomatis":"Peptostreptococcus stomatis","Gemella morbillorum":"Gemella morbillorum","Fusobacterium nucleatum":"Fusobacterium nucleatum","Parvimonas unclassified":"Parvimonas micra"}

def tests(label, exposure, subset_ids, pairs, cov_override=None):
 cv=clinical.loc[subset_ids,cov_override or covcols]; rows=[]
 for _,a in pairs.iterrows():
  s,m=a.species,a.metabolite
  if s not in exposure.columns or m not in M.columns: rows.append(dict(analysis=label,species=s,metabolite=m,n=0,r=np.nan,p=np.nan)); continue
  xr=residual(invnorm(exposure.loc[subset_ids,s]),cv); yr=residual(invnorm(M.loc[subset_ids,m]),cv); ok=np.isfinite(xr)&np.isfinite(yr)
  r,p=pearsonr(xr[ok],yr[ok]) if ok.sum()>=25 and np.std(xr[ok])>0 and np.std(yr[ok])>0 else (np.nan,np.nan)
  rows.append(dict(analysis=label,species=s,metabolite=m,n=int(ok.sum()),r=r,p=p))
 z=pd.DataFrame(rows); z["q_bh"]=bh(z.p.fillna(1)); return z

# 1) Same samples, alternative mOTU profiler. Parvimonas is explicitly a near-neighbour sensitivity only.
motu_exp=pd.DataFrame(index=ids)
for locked,observed in motu_map.items():
 if observed in motu_clr.columns: motu_exp[locked]=motu_clr[observed]
motu=tests("mOTU_CLR",motu_exp,ids,cand)
motu["exact_taxon_match"]=motu.species.ne("Parvimonas unclassified")

# 2) CRC-only association, adjusted for stage and remaining clinical covariates.
crc_ids=clinical.index[clinical.Group.isin(CRC_GROUPS)].tolist(); crc_cov=[c for c in covcols if c!="Group"]+["Stage"] if "Stage" in clinical.columns else [c for c in covcols if c!="Group"]
crc=tests("CRC_only_MetaPhlAn_CLR",meta_clr,crc_ids,cand,crc_cov)

# 3) Healthy-only association; supportive but not required.
healthy_ids=clinical.index[clinical.Group.eq("Healthy")].tolist(); healthy_cov=[c for c in covcols if c not in {"Group","Tumor location"}]
healthy=tests("Healthy_only_MetaPhlAn_CLR",meta_clr,healthy_ids,cand,healthy_cov)

# 4) Presence/absence sensitivity across the complete paired cohort.
presence=(meta_raw>0).astype(float); pres=tests("MetaPhlAn_presence_absence",presence,ids,cand)

for name,z in [("motu_locked_pairs.csv",motu),("crc_only_locked_pairs.csv",crc),("healthy_only_locked_pairs.csv",healthy),("presence_absence_locked_pairs.csv",pres)]: z.to_csv(OUT/name,index=False)

key=cand[["species","metabolite","partial_spearman_discovery","partial_spearman_validation","q_bh_family","q_bh_locked"]].copy()
for prefix,z in [("motu",motu),("crc",crc),("healthy",healthy),("presence",pres)]:
 key=key.merge(z[["species","metabolite","r","p","q_bh"]].rename(columns={"r":f"{prefix}_r","p":f"{prefix}_p","q_bh":f"{prefix}_q"}),on=["species","metabolite"],how="left")
key["motu_exact"]=key.species.ne("Parvimonas unclassified")
key["motu_concordant"]=np.sign(key.partial_spearman_discovery)==np.sign(key.motu_r)
key["crc_concordant"]=np.sign(key.partial_spearman_discovery)==np.sign(key.crc_r)
key["presence_concordant"]=np.sign(key.partial_spearman_discovery)==np.sign(key.presence_r)
key["strict_robust"]=(~key.motu_exact | (key.motu_concordant & (key.motu_q<.10))) & key.crc_concordant & (key.crc_q<.10) & key.presence_concordant & (key.presence_q<.10)
key.to_csv(OUT/"robustness_consensus.csv",index=False)
summary={"primary_reproduced_pairs":len(cand),"mOTU_exact_pairs_tested":int(motu.exact_taxon_match.sum()),"mOTU_exact_q_lt_0_10":int(((motu.q_bh<.10)&motu.exact_taxon_match).sum()),"crc_only_q_lt_0_10":int((crc.q_bh<.10).sum()),"presence_q_lt_0_10":int((pres.q_bh<.10).sum()),"strict_robust_pairs":int(key.strict_robust.sum()),"strict_pairs":key.loc[key.strict_robust,["species","metabolite","partial_spearman_discovery","partial_spearman_validation","motu_r","crc_r","presence_r"]].to_dict('records'),"parvimonas_boundary":"mOTU uses P. micra only as near-neighbour sensitivity; it is not exact validation of Parvimonas unclassified."}
(OUT/"robustness_summary.json").write_text(json.dumps(summary,indent=2),encoding="utf-8"); print(json.dumps(summary,indent=2))
